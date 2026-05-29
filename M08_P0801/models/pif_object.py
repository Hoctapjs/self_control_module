from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, AccessError, UserError
from datetime import timedelta
import os
import logging
import json
import re

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl not found, generic excel import will fail")
    openpyxl = None


class PifObject(models.Model):
    _name = 'x_psm_pif_object'
    _description = 'PIF Execution Object'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    def _get_default_tracking_steps(self):
        return [
            ('1_rsg', 'RSG creates PIF', 'RSG'),
            ('2_it_config', 'IT config', 'IT'),
            ('3_master', 'RSG setup', 'RSG'),
            ('4_lab', 'RSG lab test', 'RSG'),
            ('5_pilot', 'Pilot/Mass', 'IT'),
            ('6_val', 'PIF Valuation', 'RSG'),
            ('7_finish', 'Process Completed - Handover to Sourcing', 'System'),
        ]

    def _get_pif_size_config(self):
        return {
            'large': {
                'price_tier': 3,
                'week_number': '1st & 3rd',
                'menu_code_count': '20',
                'system_code_count': 60,
                'days_build_rfm': 10,
                'days_test_lab': 3,
                'days_backup': 2,
                'total_working_days': 15,
                'effective_rule': 'next_monday',
            },
            'medium': {
                'price_tier': 3,
                'week_number': '1st & 3rd',
                'menu_code_count': '10-20',
                'system_code_count': 45,
                'days_build_rfm': 7,
                'days_test_lab': 2,
                'days_backup': 1,
                'total_working_days': 10,
                'effective_rule': 'next_monday',
            },
            'small': {
                'price_tier': 3,
                'week_number': '1st & 3rd',
                'menu_code_count': '4-10',
                'system_code_count': 21,
                'days_build_rfm': 2,
                'days_test_lab': 1,
                'days_backup': 1,
                'total_working_days': 4,
                'effective_rule': 'next_monday',
            },
            'adhoc': {
                'price_tier': 1,
                'week_number': '2nd & 4th',
                'menu_code_count': '1-3',
                'system_code_count': 9,
                'days_build_rfm': 1,
                'days_test_lab': 1,
                'days_backup': 1,
                'total_working_days': 3,
                'effective_rule': 'after_backup',
            },
        }

    @api.model
    def _add_working_days(self, start_date, days):
        current_date = start_date
        added_days = 0
        while added_days < days:
            current_date += timedelta(days=1)
            if current_date.weekday() < 5:
                added_days += 1
        return current_date

    @api.model
    def _next_monday_on_or_after(self, target_date):
        days_until_monday = (7 - target_date.weekday()) % 7
        return target_date + timedelta(days=days_until_monday)

    @api.model
    def default_get(self, fields_list):
        defaults = super(PifObject, self).default_get(fields_list)
        if 'x_psm_process_tracking_ids' in fields_list:
            steps = self._get_default_tracking_steps()
            lines = []
            for code, name, dept in steps:
                lines.append((0, 0, {
                    'x_psm_step_code': code,
                    'x_psm_process_name': name,
                    'x_psm_department_ref': dept,
                    'x_psm_status': 'pending',
                }))
            defaults['x_psm_process_tracking_ids'] = lines
        return defaults

    name = fields.Char(string='PIF ID', required=True, copy=False, readonly=True, index=True, default=lambda self: _('PIF Creation'))

    x_psm_approval_request_id = fields.Many2one('approval.request', string='Origin Request', required=False, readonly=True)

    x_psm_request_owner_id = fields.Many2one('res.users', string='Request Owner', default=lambda self: self.env.user, readonly=True)
    x_psm_request_owner_department_id = fields.Many2one('hr.department', string='Department', compute='_compute_owner_info', store=True, readonly=True)
    x_psm_request_owner_job_id = fields.Many2one('hr.job', string='Job Position', compute='_compute_owner_info', store=True, readonly=True)

    x_psm_is_request_owner = fields.Boolean(compute='_compute_is_request_owner')
    x_psm_user_can_process_step = fields.Boolean(compute='_compute_user_can_process_step')
    x_psm_is_it_dept = fields.Boolean(compute='_compute_is_it_dept')
    x_psm_is_rsg_dept = fields.Boolean(compute='_compute_is_rsg_dept')

    x_psm_it_data_line_ids = fields.One2many(
        'x_psm_pif_data_line', 'x_psm_pif_object_id', string='IT Config Data',
        domain=[('x_psm_category', 'in', ['ssbi', 'rfm', 'pos'])],
    )

    def unlink(self):
        if not (self.env.is_superuser() or self.env.user.has_group('base.group_system')):
            raise UserError(_("You cannot delete PIF requests. Please use Refuse/Cancel instead."))
        return super(PifObject, self).unlink()

    @api.depends('state')
    def _compute_is_it_dept(self):
        it_dept = self.env.user.has_group('M08_P0801.group_pif_it') or self.env.user.has_group('M08_P0801.group_pif_manager')
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        if employee and employee.department_id and 'IT' in employee.department_id.name:
            it_dept = True
        for rec in self:
            rec.x_psm_is_it_dept = it_dept

    @api.depends('state')
    def _compute_is_rsg_dept(self):
        rsg_dept = self.env.user.has_group('M08_P0801.group_pif_rsg') or self.env.user.has_group('M08_P0801.group_pif_manager')
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        if employee and employee.department_id and 'RSG' in employee.department_id.name:
            rsg_dept = True
        for rec in self:
            rec.x_psm_is_rsg_dept = rsg_dept

    @api.depends('state')
    def _compute_user_can_process_step(self):
        state_map = {
            'rsg_create': 'RSG',
            'it_config': 'IT',
            'master_data': 'RSG',
            'lab_test': 'RSG',
            'pilot': 'IT',
            'valuation': 'RSG',
        }

        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        user_dept = employee.department_id.name if employee and employee.department_id else ''

        for rec in self:
            required = state_map.get(rec.state)

            if not required:
                rec.x_psm_user_can_process_step = False
                continue

            if required == 'RSG':
                rec.x_psm_user_can_process_step = (
                    self.env.user.has_group('M08_P0801.group_pif_rsg')
                    or self.env.user.has_group('M08_P0801.group_pif_manager')
                    or required in user_dept
                )
            elif required == 'IT':
                rec.x_psm_user_can_process_step = (
                    self.env.user.has_group('M08_P0801.group_pif_it')
                    or self.env.user.has_group('M08_P0801.group_pif_manager')
                    or required in user_dept
                )
            else:
                rec.x_psm_user_can_process_step = required in user_dept

                if self.env.is_superuser() or self.env.user.has_group('base.group_system'):
                    rec.x_psm_user_can_process_step = True

    @api.depends('x_psm_request_owner_id')
    def _compute_is_request_owner(self):
        for rec in self:
            rec.x_psm_is_request_owner = (rec.x_psm_request_owner_id == self.env.user)

    @api.depends('x_psm_request_owner_id')
    def _compute_owner_info(self):
        for rec in self:
            if rec.x_psm_request_owner_id:
                employee = self.env['hr.employee'].sudo().search(
                    [('user_id', '=', rec.x_psm_request_owner_id.id)], limit=1,
                )
                rec.x_psm_request_owner_department_id = employee.department_id if employee else False
                rec.x_psm_request_owner_job_id = employee.job_id if employee else False
            else:
                rec.x_psm_request_owner_department_id = False
                rec.x_psm_request_owner_job_id = False

    state = fields.Selection([
        ('rsg_create', 'RSG Created'),
        ('it_config', 'IT Configured'),
        ('master_data', 'Master Data Set'),
        ('lab_test', 'Lab Tested'),
        ('pilot', 'Pilot Done'),
        ('valuation', 'Valuation Done'),
        ('completed', 'Completed'),
    ], string='Status', required=True, readonly=True, copy=False, tracking=True, default='rsg_create')

    x_psm_pif_request_type = fields.Selection([
        ('menu', 'Menu'),
        ('marketing', 'Marketing'),
        ('si', 'S&I'),
        ('digital', 'Digital'),
        ('supply_chain', 'Supply Chain'),
    ], string='Request Type', default='menu')

    x_psm_program = fields.Char(string='Program')
    x_psm_initiator_dept = fields.Many2one('hr.department', string='Initiator Department')
    x_psm_new_code_seq = fields.Integer(string='New Code Required')
    x_psm_menu_item_code = fields.Char(string='Product ID / Menu Item Code', tracking=True)
    x_psm_pmix_code = fields.Char(string='Short Name / PMIX Code', tracking=True)
    x_psm_product_name_en = fields.Char(string='Product Name (EN)')
    x_psm_product_name_vi = fields.Char(string='Product Name (VI)')
    x_psm_product_form = fields.Selection([
        ('alacarte', 'A la carte'),
        ('combo', 'Combo'),
        ('meal', 'Meal'),
    ], string='Product Form')
    x_psm_menu_type = fields.Char(string='Menu Type')
    x_psm_cat_lv1 = fields.Char(string='Category LV1')
    x_psm_cat_lv2 = fields.Char(string='Category LV2')
    x_psm_cat_lv3 = fields.Char(string='Category LV3')
    x_psm_cat_lv4 = fields.Char(string='Category LV4')
    x_psm_cat_lv5 = fields.Char(string='Category LV5')
    x_psm_cat_lv6 = fields.Char(string='Category LV6')
    x_psm_cat_lv7 = fields.Char(string='Category LV7')
    x_psm_target_effective_date = fields.Date(string='Target Effective Date')
    x_psm_platform_ids = fields.Many2many(
        'x_psm_pif_platform',
        'x_psm_pif_object_platform_rel',
        'x_psm_pif_object_id',
        'x_psm_platform_id',
        string='Platforms',
    )
    x_psm_currency_id = fields.Many2one(
        'res.currency',
        string='Currency',
        default=lambda self: self.env.company.currency_id,
    )
    x_psm_price_instore = fields.Monetary(
        string='In-store/VNM Price',
        currency_field='x_psm_currency_id',
    )
    x_psm_price_delivery = fields.Monetary(
        string='Delivery Price',
        currency_field='x_psm_currency_id',
    )
    x_psm_price_tsn = fields.Monetary(
        string='TSN Price',
        currency_field='x_psm_currency_id',
    )
    x_psm_price_other = fields.Monetary(
        string='Other Price',
        currency_field='x_psm_currency_id',
    )
    x_psm_pif_size = fields.Selection([
        ('large', 'Large'),
        ('medium', 'Medium'),
        ('small', 'Small'),
        ('adhoc', 'Ad-hoc'),
    ], string='PIF Size')
    x_psm_price_tier = fields.Integer(string='Price Tier', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_week_number = fields.Char(string='Week Number', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_menu_code_count = fields.Char(string='Menu Code Count', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_system_code_count = fields.Integer(string='Total System Code', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_days_build_rfm = fields.Integer(string='Build RFM Days', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_days_test_lab = fields.Integer(string='Test LAB Days', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_days_backup = fields.Integer(string='Back up Days', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_total_working_days = fields.Integer(string='Total Working Days', compute='_compute_pif_sla', store=True, readonly=True)
    x_psm_submit_date = fields.Date(string='Submit Date', default=fields.Date.context_today)
    x_psm_effective_date_calc = fields.Date(string='Calculated Effective Date', compute='_compute_pif_sla', store=True, readonly=True)

    x_psm_pif_product_id = fields.Many2one('product.product', string='Product', required=False)
    x_psm_formula_description = fields.Text(string='Formula Description')
    x_psm_description_html = fields.Html(string='General Description')
    x_psm_supplier_info = fields.Text(string='Supplier Information')

    x_psm_pif_product_name = fields.Char(related='x_psm_pif_product_id.name', string='Product Name')
    x_psm_pif_product_code = fields.Char(related='x_psm_pif_product_id.default_code', string='Code/WRIN')
    x_psm_pif_product_vendor = fields.Char(compute='_compute_pif_product_vendor', string='Vendor')
    x_psm_pif_product_status = fields.Selection(related='x_psm_pif_product_id.x_psm_0801_pif_status', string='Product Status', readonly=True)
    x_psm_pif_bom_status = fields.Selection(related='x_psm_pif_bom_id.x_psm_0801_pif_status', string='BOM Status', readonly=True)

    x_psm_pif_bom_id = fields.Many2one('mrp.bom', string='Select BOM', domain="[('x_psm_0801_pif_status', '=', 'completed')]")
    x_psm_raw_material_ids = fields.One2many('x_psm_pif_object_raw_line', 'x_psm_pif_object_id', string='Raw Materials')
    x_psm_component_line_ids = fields.One2many('x_psm_pif_component_line', 'x_psm_pif_object_id', string='Menu Components')
    x_psm_insert_ketchup_chili = fields.Selection([
        ('yes', 'Yes'),
        ('no', 'No'),
    ], string='Insert KETCHUP & CHILI SACHET as choice?', default='no')

    x_psm_file_ssbi_html = fields.Html(string='SSBI Description')
    x_psm_file_rfm_html = fields.Html(string='RFM Description')
    x_psm_file_pos_html = fields.Html(string='POS Description')
    x_psm_img_pos = fields.Binary(string='POS Image')
    x_psm_img_pos_filename = fields.Char(string='POS Image Filename')
    x_psm_img_sok = fields.Binary(string='SOK Image')
    x_psm_img_sok_filename = fields.Char(string='SOK Image Filename')
    x_psm_img_dmb = fields.Binary(string='DMB Image')
    x_psm_img_dmb_filename = fields.Char(string='DMB Image Filename')
    x_psm_img_mop = fields.Binary(string='MOP Image')
    x_psm_img_mop_filename = fields.Char(string='MOP Image Filename')
    x_psm_setup_code_price_store = fields.Boolean(string='Code/Price/Store Setup')
    x_psm_setup_udp = fields.Boolean(string='UDP Setup')
    x_psm_setup_recipe_ssbi = fields.Boolean(string='Recipe/SSBI Setup')
    x_psm_setup_rawitem = fields.Boolean(string='Raw Item Setup')
    x_psm_display_check_ids = fields.One2many('x_psm_pif_display_check', 'x_psm_pif_object_id', string='Display Checks')

    @api.depends('x_psm_pif_size', 'x_psm_submit_date')
    def _compute_pif_sla(self):
        config_map = self._get_pif_size_config()
        for rec in self:
            config = config_map.get(rec.x_psm_pif_size)
            if not config:
                rec.x_psm_price_tier = 0
                rec.x_psm_week_number = False
                rec.x_psm_menu_code_count = False
                rec.x_psm_system_code_count = 0
                rec.x_psm_days_build_rfm = 0
                rec.x_psm_days_test_lab = 0
                rec.x_psm_days_backup = 0
                rec.x_psm_total_working_days = 0
                rec.x_psm_effective_date_calc = False
                continue

            rec.x_psm_price_tier = config['price_tier']
            rec.x_psm_week_number = config['week_number']
            rec.x_psm_menu_code_count = config['menu_code_count']
            rec.x_psm_system_code_count = config['system_code_count']
            rec.x_psm_days_build_rfm = config['days_build_rfm']
            rec.x_psm_days_test_lab = config['days_test_lab']
            rec.x_psm_days_backup = config['days_backup']
            rec.x_psm_total_working_days = config['total_working_days']

            if not rec.x_psm_submit_date:
                rec.x_psm_effective_date_calc = False
                continue

            after_working_days = rec._add_working_days(rec.x_psm_submit_date, config['total_working_days'])
            if config['effective_rule'] == 'next_monday':
                rec.x_psm_effective_date_calc = rec._next_monday_on_or_after(after_working_days)
            else:
                rec.x_psm_effective_date_calc = after_working_days

    @api.onchange('x_psm_pif_bom_id')
    def _onchange_pif_bom_id(self):
        if not self.x_psm_pif_bom_id:
            self.x_psm_raw_material_ids = [(5, 0, 0)]
            self.x_psm_component_line_ids = [(5, 0, 0)]
            return

        bom = self.x_psm_pif_bom_id
        lines = []
        component_lines = []
        for bom_line in bom.bom_line_ids:
            lines.append((0, 0, {
                'x_psm_gri_code': bom_line.product_id.default_code or '',
                'x_psm_wrin_code': bom_line.product_id.product_tmpl_id.x_psm_0801_wrin_code or '',
                'x_psm_product_id': bom_line.product_id.id,
                'x_psm_quantity': bom_line.product_qty,
                'x_psm_uom_id': bom_line.product_uom_id.id,
            }))
            component_lines.append((0, 0, {
                'x_psm_component_type': 'parent',
                'x_psm_product_id': bom_line.product_id.id,
                'x_psm_recipe_text': "%s|%s*%s" % (
                    bom_line.product_id.default_code or bom_line.product_id.display_name,
                    bom_line.product_qty,
                    bom_line.product_uom_id.name,
                ),
                'x_psm_qty': bom_line.product_qty,
            }))
        self.x_psm_raw_material_ids = [(5, 0, 0)] + lines
        self.x_psm_component_line_ids = [(5, 0, 0)] + component_lines

    @api.depends('x_psm_pif_product_id')
    def _compute_pif_product_vendor(self):
        for rec in self:
            partners = rec.x_psm_pif_product_id.seller_ids.mapped('partner_id.name')
            rec.x_psm_pif_product_vendor = ", ".join(partners) if partners else ""

    x_psm_lab_test_result = fields.Selection([
        ('pass', 'Pass'),
        ('fail', 'Fail'),
    ], string='Lab Test Result', tracking=True)
    x_psm_lab_test_note = fields.Text(string='Lab Test Notes')
    x_psm_lab_test_proof = fields.Binary(string='Lab Test Proof')
    x_psm_lab_test_proof_filename = fields.Char(string='Proof Filename')
    x_psm_lab_initiator_signoff = fields.Boolean(string='Initiator Sign-off')
    x_psm_lab_initiator_user = fields.Many2one('res.users', string='Initiator Signed By', readonly=True)
    x_psm_lab_initiator_date = fields.Datetime(string='Initiator Sign-off Date', readonly=True)
    x_psm_lab_rsg_signoff = fields.Boolean(string='RSG Sign-off')
    x_psm_lab_rsg_user = fields.Many2one('res.users', string='RSG Signed By', readonly=True)
    x_psm_lab_rsg_date = fields.Datetime(string='RSG Sign-off Date', readonly=True)

    x_psm_pilot_store_ids = fields.Many2many('res.partner', string='Target Stores / Pilot Locations', domain="[('is_company', '=', True)]")
    x_psm_pilot_note = fields.Text(string='Pilot Notes / Result')

    x_psm_data_line_ids = fields.One2many('x_psm_pif_data_line', 'x_psm_pif_object_id', string='Data Lines')
    x_psm_process_tracking_ids = fields.One2many('x_psm_pif_process_tracking', 'x_psm_pif_object_id', string='Process Tracking')
    x_psm_lab_history_ids = fields.One2many('x_psm_pif_lab_history', 'x_psm_pif_id', string='Lab History')

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('PIF Creation')) == _('PIF Creation'):
                vals['name'] = self.env['ir.sequence'].next_by_code('x_psm_pif_object') or _('PIF Creation')
        res = super(PifObject, self).create(vals_list)
        for record in res:
            if not record.x_psm_process_tracking_ids:
                record._create_default_tracking()
        return res

    @api.constrains('x_psm_pmix_code', 'x_psm_pif_request_type', 'x_psm_product_form')
    def _check_pmix_code(self):
        for rec in self:
            if not rec.x_psm_pmix_code:
                continue
            pmix_code = rec.x_psm_pmix_code.strip()
            if len(pmix_code) > 7:
                raise ValidationError(_("PMIX Code must be 7 characters or fewer."))
            if rec.x_psm_pif_request_type == 'digital' and rec.x_psm_product_form == 'combo':
                if not re.fullmatch(r'Di\d+', pmix_code):
                    raise ValidationError(_("Digital combo PMIX Code must follow the format Di + number, for example Di123."))

    @api.constrains('x_psm_menu_item_code')
    def _check_menu_item_code_unique(self):
        for rec in self:
            if not rec.x_psm_menu_item_code:
                continue
            duplicate = self.search_count([
                ('id', '!=', rec.id),
                ('x_psm_menu_item_code', '=', rec.x_psm_menu_item_code),
            ])
            if duplicate:
                raise ValidationError(_("Product ID / Menu Item Code must be unique."))

    def _create_default_tracking(self):
        steps = self._get_default_tracking_steps()
        lines = []
        for code, name, dept in steps:
            lines.append((0, 0, {
                'x_psm_step_code': code,
                'x_psm_process_name': name,
                'x_psm_department_ref': dept,
                'x_psm_status': 'pending',
            }))
        self.write({'x_psm_process_tracking_ids': lines})

    def _check_dept_permission(self, dept_name_substr):
        if self.env.is_superuser() or self.env.user.has_group('base.group_system'):
            return True
        group_map = {
            'RSG': 'M08_P0801.group_pif_rsg',
            'IT': 'M08_P0801.group_pif_it',
        }
        group_xmlid = group_map.get(dept_name_substr)
        if group_xmlid and (self.env.user.has_group(group_xmlid) or self.env.user.has_group('M08_P0801.group_pif_manager')):
            return True
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        if not employee or not employee.department_id or dept_name_substr not in employee.department_id.name:
            raise AccessError(_("You do not have permission. This step requires participation from the '%s' Department.") % dept_name_substr)
        return True

    def _complete_step(self, step_prefix):
        line = self.x_psm_process_tracking_ids.filtered(lambda l: l.x_psm_step_code.startswith(step_prefix))
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        dept_name = employee.department_id.name if employee and employee.department_id else ''
        if line:
            line.write({
                'x_psm_status': 'done',
                'x_psm_user_id': self.env.user.id,
                'x_psm_department_ref': dept_name or line.x_psm_department_ref,
                'x_psm_execution_date': fields.Datetime.now(),
            })

    def action_rsg_create(self):
        self._check_dept_permission('RSG')
        if not (self.x_psm_file_ssbi_html and self.x_psm_file_rfm_html and self.x_psm_file_pos_html):
            raise ValidationError(_("Please enter descriptions for all 3 required sections (SSBI, RFM, POS) before completing this step."))

        self.action_import_data_from_excel()
        self.write({'state': 'it_config'})
        self._complete_step('1_')

        it_dept = self.env['hr.department'].search([('name', '=', 'IT')], limit=1)
        if not it_dept:
            it_dept = self.env['hr.department'].search([('name', 'ilike', 'IT %')], limit=1)
        if it_dept and it_dept.manager_id and it_dept.manager_id.user_id:
            self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=it_dept.manager_id.user_id.id,
                summary=_('PIF Configuration Required'),
                note=_('RSG has created the PIF. Please proceed with System Configuration.'),
            )

    def action_it_config(self):
        self._check_dept_permission('IT')
        self.write({'state': 'master_data'})
        self._complete_step('2_')

        rsg_dept = self.env['hr.department'].search([('name', 'ilike', 'RSG')], limit=1)
        if rsg_dept and rsg_dept.manager_id and rsg_dept.manager_id.user_id:
            self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=rsg_dept.manager_id.user_id.id,
                summary=_('System Setup Required'),
                note=_('IT Config Done. Please proceed with Master Data Setup.'),
            )

    def action_master_setup(self):
        self._check_dept_permission('RSG')
        missing_items = []
        if not self.x_psm_setup_code_price_store:
            missing_items.append(_('Code/Price/Store Setup'))
        if not self.x_psm_setup_udp:
            missing_items.append(_('UDP Setup'))
        if not self.x_psm_setup_recipe_ssbi:
            missing_items.append(_('Recipe/SSBI Setup'))
        if not self.x_psm_setup_rawitem:
            missing_items.append(_('Raw Item Setup'))
        if missing_items:
            raise ValidationError(_("Please complete all master setup checklist items before continuing: %s") % ", ".join(missing_items))
        self.write({'state': 'lab_test'})
        self._complete_step('3_')

        rsg_dept = self.env['hr.department'].search([('name', 'ilike', 'RSG')], limit=1)
        user_to_notify = rsg_dept.manager_id.user_id.id if rsg_dept and rsg_dept.manager_id else self.env.user.id
        self.activity_schedule(
            'mail.mail_activity_data_todo',
            user_id=user_to_notify,
            summary=_('Lab Test Required'),
            note=_('Master Data Setup Done. Please proceed with Lab Testing.'),
        )

    def action_lab_initiator_signoff(self):
        for rec in self:
            if rec.state != 'lab_test':
                raise ValidationError(_("Initiator sign-off is only available during Lab Test."))
            rec.write({
                'x_psm_lab_initiator_signoff': True,
                'x_psm_lab_initiator_user': self.env.user.id,
                'x_psm_lab_initiator_date': fields.Datetime.now(),
            })

    def action_lab_rsg_signoff(self):
        self._check_dept_permission('RSG')
        for rec in self:
            if rec.state != 'lab_test':
                raise ValidationError(_("RSG sign-off is only available during Lab Test."))
            rec.write({
                'x_psm_lab_rsg_signoff': True,
                'x_psm_lab_rsg_user': self.env.user.id,
                'x_psm_lab_rsg_date': fields.Datetime.now(),
            })

    def action_finish_lab_test(self):
        self._check_dept_permission('RSG')
        if not self.x_psm_lab_test_result:
            raise ValidationError(_("Please select a Lab Test Result (Pass/Fail) before approving."))
        if self.x_psm_lab_test_result == 'pass' and not (self.x_psm_lab_initiator_signoff and self.x_psm_lab_rsg_signoff):
            raise ValidationError(_("Lab Test PASS requires both Initiator sign-off and RSG sign-off."))
        if self.x_psm_lab_test_result == 'fail' and not self.x_psm_lab_test_note:
            raise ValidationError(_("Please provide a Message/Note explaining the failure."))

        self.env['x_psm_pif_lab_history'].create({
            'x_psm_pif_id': self.id,
            'x_psm_result': self.x_psm_lab_test_result,
            'x_psm_note': self.x_psm_lab_test_note,
            'x_psm_proof_file': self.x_psm_lab_test_proof,
            'x_psm_proof_filename': self.x_psm_lab_test_proof_filename,
        })

        if self.x_psm_lab_test_result == 'pass':
            self.write({'state': 'pilot'})
            self._complete_step('4_')

            it_dept = self.env['hr.department'].search([('name', '=', 'IT')], limit=1)
            if not it_dept:
                it_dept = self.env['hr.department'].search([('name', 'ilike', 'IT %')], limit=1)
            user_to_notify = it_dept.manager_id.user_id.id if it_dept and it_dept.manager_id else self.env.user.id
            self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=user_to_notify,
                summary=_('Ready for Pilot'),
                note=_('Lab Test Passed. Please proceed with Pilot Deployment.'),
            )

        elif self.x_psm_lab_test_result == 'fail':
            self.write({
                'state': 'it_config',
                'x_psm_lab_initiator_signoff': False,
                'x_psm_lab_initiator_user': False,
                'x_psm_lab_initiator_date': False,
                'x_psm_lab_rsg_signoff': False,
                'x_psm_lab_rsg_user': False,
                'x_psm_lab_rsg_date': False,
            })

            steps_to_reset = ['2_it_config', '3_master', '4_lab', '5_pilot']
            lines_to_reset = self.x_psm_process_tracking_ids.filtered(
                lambda l: l.x_psm_step_code in steps_to_reset,
            )
            lines_to_reset.write({'x_psm_status': 'pending', 'x_psm_user_id': False, 'x_psm_execution_date': False})

            it_dept = self.env['hr.department'].search([('name', '=', 'IT')], limit=1)
            if not it_dept:
                it_dept = self.env['hr.department'].search([('name', 'ilike', 'IT %')], limit=1)
            user_to_notify = it_dept.manager_id.user_id.id if it_dept and it_dept.manager_id else self.env.user.id
            self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=user_to_notify,
                summary=_('Lab Test Failed - Revision Required'),
                note=_('Lab Test Failed. Please check notes and update IT Config.'),
            )

        self.write({
            'x_psm_lab_test_result': False,
            'x_psm_lab_test_note': False,
            'x_psm_lab_test_proof': False,
            'x_psm_lab_test_proof_filename': False,
        })

    def action_pilot_deploy(self):
        self._check_dept_permission('IT')
        self.write({'state': 'valuation'})
        self._complete_step('5_')

        rsg_dept = self.env['hr.department'].search([('name', 'ilike', 'RSG')], limit=1)
        user_to_notify = rsg_dept.manager_id.user_id.id if rsg_dept and rsg_dept.manager_id else self.env.user.id
        self.activity_schedule(
            'mail.mail_activity_data_todo',
            user_id=user_to_notify,
            summary=_('Valuation Report Required'),
            note=_('Pilot Deployment Done. Please proceed with Valuation Report.'),
        )

    def action_valuation_report(self):
        self._check_dept_permission('RSG')
        self.write({'state': 'completed'})
        self._complete_step('6_')
        self._complete_step('7_')
        self._generate_finished_wrin()

    def _generate_finished_wrin(self):
        """Generate Finished Good WRIN when PIF is approved. Format: FG-YYYY-XXXX"""
        self.ensure_one()

        product_tmpl = None
        if self.x_psm_pif_bom_id:
            product_tmpl = self.x_psm_pif_bom_id.product_tmpl_id
        elif self.x_psm_pif_product_id:
            product_tmpl = self.x_psm_pif_product_id.product_tmpl_id

        if not product_tmpl:
            return

        if not product_tmpl.x_psm_0801_wrin_code or not product_tmpl.default_code:
            new_wrin = self.env['ir.sequence'].next_by_code('product.wrin.finished') or \
                       f"FG-{fields.Date.today().year}-{product_tmpl.id:04d}"

            product_tmpl.sudo().write({
                'x_psm_0801_wrin_code': new_wrin,
                'default_code': new_wrin,
                'x_psm_0801_pif_status': 'completed',
            })

            if self.x_psm_pif_bom_id:
                self.x_psm_pif_bom_id.sudo().write({'x_psm_0801_pif_status': 'completed'})

            self.message_post(
                body=f"WRIN Thành phẩm đã được tạo: {new_wrin}",
                message_type='notification',
            )

    def action_import_data_from_excel(self):
        """Read PIF Excel from configured folder or file/ folder and populate data lines."""
        file_dir = self.env['ir.config_parameter'].sudo().get_param('m08_p0801.pif_import_dir', '/mnt/file')
        if not os.path.exists(file_dir):
            base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            file_dir = os.path.join(base_path, 'file')
        target_file = None

        if os.path.exists(file_dir):
            for fname in os.listdir(file_dir):
                if fname.endswith('.xlsx') and 'PIF' in fname:
                    target_file = os.path.join(file_dir, fname)
                    break

        if not target_file or not os.path.exists(target_file):
            self.env['bus.bus']._sendone(self.env.user.partner_id, 'simple_notification', {
                'title': _('File Not Found'),
                'message': _('Could not find the PIF Excel file in file/ folder.'),
                'type': 'danger',
            })
            return

        if not openpyxl:
            return

        wb = openpyxl.load_workbook(target_file, data_only=True, read_only=True)

        data_vals = []
        sheet_map = {
            'SSBI': 'ssbi',
            'RFM': 'rfm',
            'RFM 2508': 'rfm',
            'RFM 0909': 'rfm',
            'POS layout': 'pos',
            'POS layout 0909': 'pos',
            'POS layout 08112024': 'pos',
        }

        for sheet_name in wb.sheetnames:
            category = sheet_map.get(sheet_name)
            if not category:
                if 'SSBI' in sheet_name: category = 'ssbi'
                elif 'RFM' in sheet_name: category = 'rfm'
                elif 'POS' in sheet_name: category = 'pos'
                elif 'MDS' in sheet_name: category = 'mds'
                elif 'UDP' in sheet_name: category = 'udp'

            if category:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = rows[0]
                for i, row in enumerate(rows[1:], start=2):
                    for j, cell_value in enumerate(row):
                        if cell_value:
                            key = f"{headers[j] if j < len(headers) and headers[j] else 'Col'+str(j)} (R{i})"
                            data_vals.append({
                                'x_psm_category': category,
                                'x_psm_key': key,
                                'x_psm_value': str(cell_value),
                            })

        if data_vals:
            self.x_psm_data_line_ids = [(0, 0, v) for v in data_vals]


class PifPlatform(models.Model):
    _name = 'x_psm_pif_platform'
    _description = 'PIF Platform'
    _order = 'sequence, name'

    name = fields.Char(string='Platform', required=True)
    code = fields.Char(string='Code', required=True)
    sequence = fields.Integer(default=10)
    active = fields.Boolean(default=True)

    _sql_constraints = [
        ('code_unique', 'unique(code)', 'Platform code must be unique.'),
    ]


class PifComponentLine(models.Model):
    _name = 'x_psm_pif_component_line'
    _description = 'PIF Menu Component Line'
    _order = 'x_psm_component_type, id'

    x_psm_pif_object_id = fields.Many2one('x_psm_pif_object', string='PIF Object', ondelete='cascade')
    x_psm_component_type = fields.Selection([
        ('parent', 'Parent'),
        ('bic_flavor', 'BIC Flavor'),
        ('side_dish', 'Side Dish'),
        ('side_dish_special', 'Side Dish Special'),
        ('drink', 'Drink'),
        ('drink_special', 'Drink Special'),
        ('other', 'Other'),
    ], string='Component Type', required=True, default='parent')
    x_psm_product_id = fields.Many2one('product.product', string='Item Code')
    x_psm_recipe_text = fields.Text(string='Item Parent Recipe')
    x_psm_qty = fields.Float(string='QTY')
    x_psm_price_vnm = fields.Float(string='VNM Price')
    x_psm_price_delivery = fields.Float(string='Delivery Price')
    x_psm_price_tsn = fields.Float(string='TSN Price')
    x_psm_is_default = fields.Boolean(string='Default')
    x_psm_default_size = fields.Char(string='Default Size')
    x_psm_allow_upsize = fields.Boolean(string='Allow Upsize?')
    x_psm_allow_upgrade = fields.Boolean(string='Allow Upgrade?')
    x_psm_choice_group = fields.Char(string='Choice Group')


class PifDisplayCheck(models.Model):
    _name = 'x_psm_pif_display_check'
    _description = 'PIF Display Check'
    _order = 'x_psm_check_date desc, id desc'

    x_psm_pif_object_id = fields.Many2one('x_psm_pif_object', string='PIF Object', ondelete='cascade')
    x_psm_department = fields.Selection([
        ('menu', 'Menu'),
        ('marketing', 'Marketing'),
        ('si', 'S&I'),
        ('digital', 'Digital'),
        ('supply_chain', 'Supply Chain'),
    ], string='Department', required=True)
    x_psm_checked_by = fields.Many2one('res.users', string='Checked By', default=lambda self: self.env.user)
    x_psm_check_date = fields.Datetime(string='Check Date', default=fields.Datetime.now)
    x_psm_status = fields.Selection([
        ('pass', 'Pass'),
        ('fail', 'Fail'),
    ], string='Status', default='pass', required=True)
    x_psm_note = fields.Text(string='Note')


class PifDataLine(models.Model):
    _name = 'x_psm_pif_data_line'
    _description = 'PIF Command Data Line'

    x_psm_pif_object_id = fields.Many2one('x_psm_pif_object', string='PIF Object', ondelete='cascade')
    x_psm_category = fields.Selection([
        ('ssbi', 'SSBI'),
        ('rfm', 'RFM'),
        ('pos', 'POS'),
        ('mds', 'MDS'),
        ('udp', 'UDP'),
        ('access', 'Access & Audit'),
        ('other', 'Other'),
    ], string='System/Category', required=True, default='other')
    x_psm_key = fields.Char(string='Command/Parameter', required=True)
    x_psm_value = fields.Text(string='Value/Config')


class PifObjectRawLine(models.Model):
    _name = 'x_psm_pif_object_raw_line'
    _description = 'PIF Object Raw Material'

    x_psm_pif_object_id = fields.Many2one('x_psm_pif_object', string='PIF Object', ondelete='cascade')

    x_psm_gri_code = fields.Char(string='GRI')
    x_psm_wrin_code = fields.Char(string='WRIN')
    x_psm_product_id = fields.Many2one('product.product', string='Raw Item Name')
    x_psm_quantity = fields.Float(string='Quantity')
    x_psm_uom_id = fields.Many2one('uom.uom', string='UoM')


class PifProcessTracking(models.Model):
    _name = 'x_psm_pif_process_tracking'
    _description = 'PIF Process Tracking Step'
    _order = 'id'

    x_psm_pif_object_id = fields.Many2one('x_psm_pif_object', string='PIF Object', ondelete='cascade')
    x_psm_step_code = fields.Char(string='Step Code', required=True)

    x_psm_process_name = fields.Char(string='Process Name')
    x_psm_department_ref = fields.Char(string='Department')

    x_psm_status = fields.Selection([
        ('pending', 'Pending'),
        ('processing', 'Processing'),
        ('done', 'Done'),
        ('fail', 'Failed'),
    ], string='Status', default='pending')

    x_psm_user_id = fields.Many2one('res.users', string='Executed By')
    x_psm_execution_date = fields.Datetime(string='Execution Date')


class PifLabHistory(models.Model):
    _name = 'x_psm_pif_lab_history'
    _description = 'PIF Lab Test History'
    _order = 'x_psm_test_date desc'

    x_psm_pif_id = fields.Many2one('x_psm_pif_object', string='PIF', ondelete='cascade')
    x_psm_test_date = fields.Datetime(string='Test Date', default=fields.Datetime.now, readonly=True)
    x_psm_tester_id = fields.Many2one('res.users', string='Tester', default=lambda self: self.env.user, readonly=True)
    x_psm_result = fields.Selection([
        ('pass', 'Pass'),
        ('fail', 'Fail'),
    ], string='Result', readonly=True)
    x_psm_note = fields.Text(string='Note', readonly=True)
    x_psm_proof_file = fields.Binary(string='Attachment', readonly=True)
    x_psm_proof_filename = fields.Char(string='Filename', readonly=True)
